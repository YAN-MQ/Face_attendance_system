from flask import Flask, request, jsonify, render_template, send_from_directory
from flask_cors import CORS
import os
import cv2
import numpy as np
import json
import base64
from datetime import datetime
# 修改这一行导入
from flask.json import provider
import traceback

# 导入Excel相关模块
import pandas as pd
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename

# 导入自定义活体检测模块
from models.blink_detection import BlinkDetector
from models.deep_learning_liveness import DeepLearningLiveness
from models.api_liveness import APILiveness
from utils import to_serializable, ensure_directory

# 导入数据库相关模块
from db_config import SessionLocal, engine, Base
import db_utils
from models.database_models import Student, FaceImage, AttendanceRecord, Class
from models.face_recognition_utils import FaceRecognitionUtils
from models.liveness_detection_improved import ImprovedLivenessDetection

app = Flask(__name__, static_folder='../static', static_url_path='/static')
CORS(app)  # 启用跨域支持

# 创建自定义 JSON 编码器来处理 NumPy 类型
class CustomJSONEncoder(provider.DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()  # 将 ndarray 转换为 list
        if isinstance(obj, np.bool_):
            return bool(obj)     # 将 numpy.bool_ 转换为 Python bool
        if isinstance(obj, np.integer):
            return int(obj)      # 将 numpy.integer 转换为 Python int
        if isinstance(obj, np.floating):
            return float(obj)    # 将 numpy.floating 转换为 Python float
        if isinstance(obj, np.str_):
            return str(obj)      # 将 numpy.str_ 转换为 Python str
        return super().default(obj)

# 将自定义编码器应用到 Flask 应用
app.json_provider_class = CustomJSONEncoder
app.json = CustomJSONEncoder(app)

# 初始化各种活体检测器
blink_detector = BlinkDetector()
deep_learning_detector = DeepLearningLiveness()
api_detector = APILiveness()
# 添加改进的活体检测器
improved_detector = ImprovedLivenessDetection()

# 初始化人脸识别工具
face_recognition_utils = FaceRecognitionUtils(face_db_dir="static/face_db")

# 确保人脸数据库目录存在
FACE_DB_DIR = "static/face_db"
ensure_directory(FACE_DB_DIR)

# 获取数据库会话
def get_db():
    db = SessionLocal()
    try:
        return db
    finally:
        pass  # 这里不关闭会话，由调用者负责关闭

# 为确保会话正确关闭的上下文管理器
def with_db_session(func):
    def wrapper(*args, **kwargs):
        db = get_db()
        try:
            result = func(*args, db=db, **kwargs)
            return result
        except Exception as e:
            db.rollback()
            raise e
        finally:
            db.close()
    return wrapper

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/detect_liveness', methods=['POST'])
def detect_liveness():
    if 'image' not in request.json or 'method' not in request.json:
        return jsonify({"error": "Missing image or method"}), 400
    
    # 解码BASE64图像
    image_data = request.json['image'].split(',')[1]
    image_bytes = base64.b64decode(image_data)
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    
    method = request.json['method']
    result = {"is_live": False, "message": "未知错误", "confidence": 0}
    
    # 根据选择的方法进行活体检测
    if method == "blink":
        result = blink_detector.detect(img)
    elif method == "deep_learning":
        result = deep_learning_detector.detect(img)
    elif method == "api":
        result = api_detector.detect(img)
    elif method == "improved":
        # 使用改进的活体检测
        result = improved_detector.detect(img)
    else:
        return jsonify({"error": "不支持的活体检测方法"}), 400
    
    # 在返回前转换结果
    result = to_serializable(result)
    return jsonify(result)

@app.route('/api/recognize_face', methods=['POST'])
def recognize_face():
    if 'image' not in request.json:
        return jsonify({"error": "缺少图像数据"}), 400
    
    # 解码BASE64图像
    image_data = request.json['image'].split(',')[1]
    image_bytes = base64.b64decode(image_data)
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    
    # 检查是否是活体
    liveness_result = None
    if 'method' in request.json:
        method = request.json['method']
        
        if method == "blink":
            liveness_result = blink_detector.detect(img)
        elif method == "deep_learning":
            liveness_result = deep_learning_detector.detect(img)
        elif method == "api":
            liveness_result = api_detector.detect(img)
        elif method == "improved":
            # 使用改进的活体检测
            liveness_result = improved_detector.detect(img)
            
        if not liveness_result or not liveness_result.get("is_live", False):
            return jsonify({
                "success": False, 
                "message": "活体检测失败", 
                "liveness_result": liveness_result
            })
    
    # 进行人脸识别
    recognition_result = face_recognition_utils.recognize_face(img)
    
    student_id = recognition_result.get("student_id", "未识别")
    student_name = "未知"
    similarity = recognition_result.get("similarity", 0.0)
    
    # 如果识别成功，获取学生姓名
    if recognition_result.get("success", False):
        db = None
        try:
            db = get_db()
            student = db_utils.get_student_by_id(db, student_id)
            if student:
                student_name = student.name
        except Exception as e:
            print(f"获取学生信息出错: {str(e)}")
        finally:
            if db:
                db.close()
    
    # 模拟识别结果
    success = recognition_result.get("success", False)
    
    # 记录考勤
    if success:
        db = None
        try:
            db = get_db()
            db_utils.record_attendance(
                db=db,
                student_id=student_id,
                recognition_confidence=similarity,
                liveness_method=request.json.get('method'),
                liveness_confidence=liveness_result.get('confidence') if liveness_result else None
            )
        except Exception as e:
            if db:
                db.rollback()
            print(f"记录考勤出错: {str(e)}")
        finally:
            if db:
                db.close()
    
    return jsonify({
        "success": success,
        "student_id": student_id,
        "student_name": student_name,
        "similarity": similarity,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

@app.route('/api/register_face', methods=['POST'])
def register_face():
    print("接收到register_face请求")
    
    # 验证请求参数
    if not all(k in request.json for k in ['image', 'student_id', 'student_name']):
        error_msg = "缺少必要参数"
        print(f"错误: {error_msg}")
        return jsonify({"error": error_msg, "success": False}), 400
    
    student_id = request.json['student_id']
    student_name = request.json['student_name']
    class_id = request.json.get('class_id')
    
    print(f"处理注册请求: 学生ID={student_id}, 姓名={student_name}, 班级ID={class_id}")
    
    # 解码BASE64图像
    try:
        print("开始解码图像...")
        image_data = request.json['image']
        
        # 检查并去除数据URL前缀
        if ',' in image_data:
            print("检测到数据URL前缀，处理中...")
            # 提取MIME类型和Base64部分
            header, base64_data = image_data.split(',', 1)
            print(f"图像头部: {header[:30]}...")
        else:
            # 没有前缀，直接使用
            base64_data = image_data
            print("图像没有数据URL前缀")
            
        print(f"Base64数据长度: {len(base64_data)}")
        print(f"Base64数据前几个字符: {base64_data[:20]}...")
        
        # 解码Base64数据
        try:
            image_bytes = base64.b64decode(base64_data)
            print(f"解码后的图像字节长度: {len(image_bytes)}")
        except Exception as e:
            error_msg = f"Base64解码失败: {str(e)}"
            print(f"错误: {error_msg}")
            traceback.print_exc()
            return jsonify({"error": error_msg, "success": False}), 400
        
        # 解码图像数据
        try:
            nparr = np.frombuffer(image_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if img is None:
                error_msg = "图像解码失败，可能是格式不支持或数据损坏"
                print(f"错误: {error_msg}")
                
                # 保存原始字节用于调试
                debug_path = os.path.join("static/uploads", f"debug_{student_id}.bin")
                os.makedirs(os.path.dirname(debug_path), exist_ok=True)
                with open(debug_path, "wb") as f:
                    f.write(image_bytes)
                print(f"已保存调试数据到: {debug_path}")
                
                return jsonify({"error": error_msg, "success": False}), 400
                
            print(f"图像解码成功: 尺寸={img.shape}")
        except Exception as e:
            error_msg = f"图像处理出错: {str(e)}"
            print(f"错误: {error_msg}")
            traceback.print_exc()
            return jsonify({"error": error_msg, "success": False}), 400
    except Exception as e:
        error_msg = f"图像处理出错: {str(e)}"
        print(f"错误: {error_msg}")
        traceback.print_exc()
        return jsonify({"error": error_msg, "success": False}), 400
    
    # 图像解码后，添加人脸编码到缓存
    try:
        # 将人脸编码添加到缓存
        face_encoding_result = face_recognition_utils.add_face_encoding(student_id, img)
        
        if not face_encoding_result.get("success", False):
            print(f"添加人脸编码失败: {face_encoding_result.get('message')}")
            # 继续处理，因为我们仍然会保存原始图像
    except Exception as e:
        print(f"处理人脸编码时出错: {str(e)}")
        # 继续处理，保存原始图像
    
    db = None
    try:
        # 获取数据库会话
        print("获取数据库会话...")
        db = get_db()
        
        # 使用os.makedirs确保目录存在，设置权限
        student_dir = os.path.join(FACE_DB_DIR, student_id)
        print(f"确保目录存在: {student_dir}")
        try:
            os.makedirs(student_dir, exist_ok=True)
            print(f"目录已创建或已存在: {student_dir}")
        except Exception as e:
            error_msg = f"创建目录失败: {str(e)}"
            print(f"错误: {error_msg}")
            traceback.print_exc()
            return jsonify({"error": error_msg, "success": False}), 500
        
        # 保存图像
        image_filename = f"{student_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
        image_path = os.path.join(student_dir, image_filename)
        
        # 使用绝对路径保存图片
        abs_image_path = os.path.abspath(image_path)
        print(f"保存图像到: {abs_image_path}")
        
        try:
            # 检查目录是否可写
            if not os.access(os.path.dirname(abs_image_path), os.W_OK):
                error_msg = f"目录没有写入权限: {os.path.dirname(abs_image_path)}"
                print(f"错误: {error_msg}")
                return jsonify({"error": error_msg, "success": False}), 500
            
            # 尝试保存临时文件测试写入权限
            test_file = os.path.join(student_dir, "test_write.txt")
            try:
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                print("写入权限测试成功")
            except Exception as e:
                error_msg = f"写入测试失败: {str(e)}"
                print(f"错误: {error_msg}")
                return jsonify({"error": error_msg, "success": False}), 500
            
            # 尝试用PIL保存图像（替代方法）
            if img is not None:
                from PIL import Image
                import io
                try:
                    # 转换OpenCV图像为PIL图像
                    cv_img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    pil_img = Image.fromarray(cv_img_rgb)
                    pil_img.save(abs_image_path)
                    print(f"使用PIL保存图像成功")
                    success = True
                except Exception as e:
                    # 如果PIL保存失败，尝试使用OpenCV保存
                    print(f"PIL保存失败: {str(e)}，尝试使用OpenCV...")
                    success = cv2.imwrite(abs_image_path, img)
            else:
                success = False
                
            if not success:
                error_msg = "图像保存失败，请检查文件路径和权限"
                print(f"错误: {error_msg}")
                return jsonify({"error": error_msg, "success": False}), 500
            
            print(f"图像保存成功: {abs_image_path}")
            
            # 检查文件是否实际存在
            if not os.path.exists(abs_image_path):
                error_msg = f"图像保存失败，保存后文件不存在: {abs_image_path}"
                print(f"错误: {error_msg}")
                return jsonify({"error": error_msg, "success": False}), 500
                
            print(f"图像保存成功，文件大小: {os.path.getsize(abs_image_path)} 字节")
        except Exception as e:
            error_msg = f"图像保存过程出错: {str(e)}"
            print(f"错误: {error_msg}")
            traceback.print_exc()
            return jsonify({"error": error_msg, "success": False}), 500
        
        # 检查班级是否存在
        if class_id:
            class_obj = db.query(Class).filter(Class.id == class_id).first()
            if not class_obj:
                print(f"警告: 班级ID {class_id} 不存在，尝试创建默认班级...")
                # 尝试创建班级
                class_obj = Class(
                    id=class_id,
                    name=f"班级 {class_id}",
                    description="自动创建的班级"
                )
                db.add(class_obj)
                db.commit()
                print(f"已创建班级: {class_id}")
        
        # 保存学生信息到数据库
        print(f"检查学生是否存在: {student_id}")
        db_student = db_utils.get_student_by_id(db, student_id)
        if not db_student:
            print(f"创建新学生: {student_id}")
            db_student = db_utils.create_student(db, student_id, student_name, class_id)
            print(f"学生创建成功: {db_student.id}")
        else:
            print(f"学生已存在: {db_student.id}")
        
        # 保存人脸图像信息到数据库
        print(f"保存人脸图像记录...")
        db_face = db_utils.add_face_image(db, student_id, abs_image_path)
        print(f"人脸图像记录保存成功: ID={db_face.id}")
        
        print(f"注册成功: 学生={student_name}({student_id})")
        return jsonify({
            "success": True,
            "message": f"成功注册学生 {student_name}({student_id}) 的人脸信息",
            "saved_path": abs_image_path
        })
    except Exception as e:
        error_msg = f"保存过程出错: {str(e)}"
        print(f"错误: {error_msg}")
        traceback.print_exc()
        if db:
            db.rollback()
        return jsonify({"error": error_msg, "success": False}), 500
    finally:
        if db:
            db.close()
            print("数据库连接已关闭")

@app.route('/api/get_attendance', methods=['GET'])
def get_attendance():
    db = None
    try:
        db = get_db()
        attendance_records = db_utils.get_all_attendance_records(db)
        
        # 将ORM对象转换为JSON可序列化的字典
        result = []
        for record in attendance_records:
            student = db_utils.get_student_by_id(db, record.student_id)
            result.append({
                "id": record.id,
                "student_id": record.student_id,
                "student_name": student.name if student else "未知",
                "date": record.date.strftime("%Y-%m-%d"),
                "time": record.time.strftime("%H:%M:%S"),
                "status": record.status,
                "recognition_confidence": record.recognition_confidence,
                "liveness_method": record.liveness_method,
                "liveness_confidence": record.liveness_confidence
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"获取考勤记录失败: {str(e)}"}), 500
    finally:
        if db:
            db.close()

@app.route('/api/class_info', methods=['GET', 'POST'])
def class_info():
    db = None
    try:
        db = get_db()
        
        if request.method == 'GET':
            try:
                classes = db_utils.get_all_classes(db)
                
                # 将ORM对象转换为JSON可序列化的字典
                result = {"classes": []}
                for cls in classes:
                    class_dict = {
                        "id": cls.id,
                        "name": cls.name,
                        "teacher": cls.teacher,
                        "description": cls.description,
                        "students": []
                    }
                    
                    # 添加班级学生
                    for student in cls.students:
                        class_dict["students"].append({
                            "id": student.id,
                            "name": student.name
                        })
                    
                    result["classes"].append(class_dict)
                
                return jsonify(result)
            except Exception as e:
                return jsonify({"error": f"获取班级信息失败: {str(e)}"}), 500
        
        elif request.method == 'POST':
            try:
                new_class = request.json
                
                # 创建新班级
                db_class = db_utils.create_class(
                    db=db,
                    class_id=new_class["id"],
                    name=new_class["name"],
                    teacher=new_class.get("teacher"),
                    description=new_class.get("description")
                )
                
                return jsonify({"success": True, "message": "班级信息添加成功"})
            except Exception as e:
                db.rollback()
                return jsonify({"error": f"添加班级信息失败: {str(e)}"}), 500
    finally:
        if db:
            db.close()

@app.route('/api/livenessdetection_methods', methods=['GET'])
def livenessdetection_methods():
    """获取可用的活体检测方法"""
    methods = [
        {"id": "blink", "name": "眨眼检测", "description": "基于眨眼动作的活体检测"},
        {"id": "deep_learning", "name": "深度学习", "description": "基于深度学习的纹理分析活体检测"},
        {"id": "api", "name": "第三方API", "description": "调用腾讯云人脸识别API进行活体检测"},
        {"id": "improved", "name": "改进多模态", "description": "综合多种技术的改进活体检测"}
    ]
    
    return jsonify(methods)

@app.route('/api/student_template', methods=['GET'])
def student_template():
    """提供学生信息模板下载"""
    template_path = os.path.join(app.static_folder, 'templates')
    return send_from_directory(template_path, '学生信息模板.xlsx', as_attachment=True)

@app.route('/api/generate_excel_template', methods=['GET', 'OPTIONS'])
def generate_excel_template():
    """生成Excel导入模板"""
    # 处理预检请求
    if request.method == 'OPTIONS':
        response = app.make_response("")
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Methods', 'GET, OPTIONS')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Max-Age', '86400')
        return response
        
    try:
        # 获取班级信息用于下拉列表
        db = get_db()
        classes = db_utils.get_all_classes(db)
        class_list = [{"id": cls.id, "name": cls.name} for cls in classes]
        
        # 创建一个新的Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生人脸注册信息"
        
        # 定义样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        
        # 添加表头
        headers = ["学号(*)", "姓名(*)", "班级ID(*)", "图片路径(*)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = border
        
        # 添加一些示例数据
        sample_data = [
            ["10001", "张三", class_list[0]["id"] if class_list else "class1", "C:/Photos/10001.jpg"],
            ["10002", "李四", class_list[0]["id"] if class_list else "class1", "C:/Photos/10002.jpg"]
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
        
        # 添加数据有效性验证（班级ID下拉列表）
        if class_list:
            class_ids = [cls["id"] for cls in class_list]
            data_validation = openpyxl.worksheet.datavalidation.DataValidation(
                type="list", 
                formula1=f'"{",".join(class_ids)}"',
                allow_blank=True
            )
            data_validation.add(f'C2:C1000')  # 应用到C列
            ws.add_data_validation(data_validation)
        
        # 添加说明工作表
        ws_help = wb.create_sheet(title="填表说明")
        ws_help.column_dimensions['A'].width = 80
        
        instructions = [
            "填表说明：",
            "1. 带(*)的列为必填项",
            "2. 学号：必须是唯一的学生ID",
            "3. 姓名：学生姓名",
            "4. 班级ID：从下拉列表中选择班级ID",
            "5. 图片路径：学生照片的完整路径，支持JPG、PNG格式",
            "",
            "注意事项：",
            "1. 请确保照片清晰、正面且光线充足",
            "2. 导入时会自动处理照片并注册到系统中",
            "3. 若学生已存在，将更新对应信息"
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            cell = ws_help.cell(row=row_num, column=1)
            cell.value = instruction
        
        # 添加班级ID映射表
        if class_list:
            ws_help.cell(row=len(instructions) + 2, column=1).value = "班级ID参考表："
            ws_help.cell(row=len(instructions) + 3, column=1).value = "班级ID\t班级名称"
            
            for i, cls in enumerate(class_list):
                ws_help.cell(row=len(instructions) + 4 + i, column=1).value = f"{cls['id']}\t{cls['name']}"
        
        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        # 读取文件并设置响应
        with open(temp_file.name, 'rb') as f:
            data = f.read()
        
        # 删除临时文件
        os.unlink(temp_file.name)
        
        # 设置响应头，让浏览器下载文件
        response = app.make_response(data)
        response.headers["Content-Disposition"] = "attachment; filename=学生人脸注册模板.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # 添加CORS头，确保跨域请求能正确处理文件下载
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type"
        response.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
        response.headers["Access-Control-Expose-Headers"] = "Content-Disposition"
        
        return response
        
    except Exception as e:
        print(f"生成Excel模板出错: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": f"生成Excel模板失败: {str(e)}"}), 500
    finally:
        if 'db' in locals():
            db.close()

@app.route('/api/import_excel', methods=['POST'])
def import_excel():
    """从Excel文件导入学生信息并注册"""
    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件", "success": False}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "文件名为空", "success": False}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "只支持Excel文件(.xlsx, .xls)", "success": False}), 400
    
    # 保存上传的文件
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, secure_filename(file.filename))
    file.save(file_path)
    
    db = None
    result = {
        "success": True,
        "total": 0,
        "success_count": 0,
        "failed_count": 0,
        "details": []
    }
    
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path, sheet_name=0)
        required_cols = ["学号(*)", "姓名(*)", "班级ID(*)", "图片路径(*)"]
        
        # 检查是否包含所有必要列
        for col in required_cols:
            if col not in df.columns:
                return jsonify({"error": f"Excel文件缺少必要列: {col}", "success": False}), 400
        
        # 重命名列名为英文，方便处理
        df.columns = [
            'student_id' if col == "学号(*)" else
            'student_name' if col == "姓名(*)" else
            'class_id' if col == "班级ID(*)" else
            'image_path' if col == "图片路径(*)" else
            col for col in df.columns
        ]
        
        db = get_db()
        result["total"] = len(df)
        
        # 遍历每一行，处理学生信息
        for index, row in df.iterrows():
            # 提取每行的数据
            student_id = str(row['student_id'])
            student_name = str(row['student_name'])
            class_id = str(row['class_id'])
            image_path = str(row['image_path'])
            
            record = {
                "student_id": student_id,
                "student_name": student_name,
                "class_id": class_id,
                "image_path": image_path,
                "success": False,
                "message": ""
            }
            
            # 验证数据
            if pd.isna(row['student_id']) or pd.isna(row['student_name']) or pd.isna(row['class_id']) or pd.isna(row['image_path']):
                record["message"] = "数据不完整，存在空值"
                result["details"].append(record)
                result["failed_count"] += 1
                continue
            
            # 检查图片文件是否存在
            if not os.path.exists(image_path):
                record["message"] = f"图片路径不存在: {image_path}"
                result["details"].append(record)
                result["failed_count"] += 1
                continue
            
            try:
                # 读取图片
                img = cv2.imread(image_path)
                if img is None:
                    record["message"] = f"无法读取图片: {image_path}"
                    result["details"].append(record)
                    result["failed_count"] += 1
                    continue
                
                # 检查班级是否存在，不存在则创建
                class_obj = db.query(Class).filter(Class.id == class_id).first()
                if not class_obj:
                    class_obj = Class(
                        id=class_id,
                        name=f"班级 {class_id}",
                        description="Excel导入自动创建的班级"
                    )
                    db.add(class_obj)
                    db.commit()
                
                # 检查学生是否存在
                db_student = db_utils.get_student_by_id(db, student_id)
                if not db_student:
                    db_student = db_utils.create_student(db, student_id, student_name, class_id)
                
                # 保存人脸图像
                student_dir = os.path.join(FACE_DB_DIR, student_id)
                ensure_directory(student_dir)
                
                # 生成目标图像路径
                image_filename = f"{student_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
                target_image_path = os.path.join(student_dir, image_filename)
                
                # 保存图像到人脸数据库
                cv2.imwrite(target_image_path, img)
                
                # 将图像添加到人脸编码缓存
                face_encoding_result = face_recognition_utils.add_face_encoding(student_id, img)
                
                # 保存人脸图像信息到数据库
                db_face = db_utils.add_face_image(db, student_id, target_image_path)
                
                record["success"] = True
                record["message"] = "注册成功"
                result["success_count"] += 1
            except Exception as e:
                record["message"] = f"处理出错: {str(e)}"
                result["failed_count"] += 1
                traceback.print_exc()
            
            result["details"].append(record)
        
        return jsonify(result)
    except Exception as e:
        print(f"Excel导入出错: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "error": f"Excel导入失败: {str(e)}",
            "success": False
        }), 500
    finally:
        # 清理临时文件
        if os.path.exists(file_path):
            os.unlink(file_path)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)
        # 关闭数据库连接
        if db:
            db.close()

if __name__ == '__main__':
    # 确保数据库表存在
    Base.metadata.create_all(bind=engine)
    
    app.run(debug=True, host='0.0.0.0', port=5000)